"""
Tabular data parser for SuperInsight Platform.

Provides CSV and Excel file parsing using pandas, returning unified TabularData.
"""

from __future__ import annotations

import logging
import os
from dataclasses import dataclass
from typing import Any, Optional

logger = logging.getLogger(__name__)

# Optional dependencies
try:
    import pandas as pd
    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False

try:
    import openpyxl  # noqa: F401
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

MAX_FILE_SIZE_BYTES = 100 * 1024 * 1024  # 100MB
VALID_FILE_TYPES = {"csv", "excel"}


@dataclass
class TabularData:
    """Unified result from CSV/Excel parsing."""
    headers: list[str]
    rows: list[dict[str, Any]]
    row_count: int
    file_type: str  # "csv" | "excel"
    sheet_name: str | None = None


class TabularParser:
    """Parser for CSV and Excel files using pandas."""

    def parse(self, file_path: str, file_type: str) -> TabularData:
        """Parse a tabular file and return unified TabularData.

        Args:
            file_path: Path to the file on disk.
            file_type: Either "csv" or "excel".

        Returns:
            TabularData with headers, rows, and metadata.

        Raises:
            ValueError: If file_type is invalid or file fails validation.
            FileNotFoundError: If file_path does not exist.
            RuntimeError: If parsing fails due to corruption or encoding issues.
        """
        if file_type not in VALID_FILE_TYPES:
            raise ValueError(
                f"Unsupported file_type: '{file_type}'. Must be one of {VALID_FILE_TYPES}"
            )

        self._validate_file(file_path)

        if file_type == "csv":
            return self._parse_csv(file_path)
        return self._parse_excel(file_path)

    # ------------------------------------------------------------------
    # Internal helpers
    # ------------------------------------------------------------------

    def _validate_file(self, file_path: str) -> None:
        """Validate file existence and size."""
        if not os.path.isfile(file_path):
            raise FileNotFoundError(f"File not found: {file_path}")

        file_size = os.path.getsize(file_path)
        if file_size > MAX_FILE_SIZE_BYTES:
            raise ValueError(
                f"File size ({file_size / 1024 / 1024:.1f}MB) exceeds "
                f"limit of {MAX_FILE_SIZE_BYTES / 1024 / 1024:.0f}MB"
            )
        if file_size == 0:
            raise ValueError("File is empty")

    def _parse_csv(self, file_path: str) -> TabularData:
        """Parse a CSV file using pandas.read_csv."""
        if not PANDAS_AVAILABLE:
            raise RuntimeError(
                "pandas is required for CSV parsing. Install with: pip install pandas"
            )

        try:
            df = pd.read_csv(file_path, encoding="utf-8")
        except UnicodeDecodeError:
            logger.warning("UTF-8 decode failed for %s, retrying with latin-1", file_path)
            try:
                df = pd.read_csv(file_path, encoding="latin-1")
            except Exception as exc:
                raise RuntimeError(f"Failed to parse CSV file: {exc}") from exc
        except Exception as exc:
            raise RuntimeError(f"Failed to parse CSV file: {exc}") from exc

        return self._dataframe_to_tabular(df, file_type="csv")

    def _parse_excel(self, file_path: str) -> TabularData:
        """Parse an Excel file using pandas.read_excel (openpyxl engine)."""
        if not PANDAS_AVAILABLE:
            raise RuntimeError(
                "pandas is required for Excel parsing. Install with: pip install pandas"
            )
        if not OPENPYXL_AVAILABLE:
            raise RuntimeError(
                "openpyxl is required for Excel parsing. Install with: pip install openpyxl"
            )

        try:
            df = pd.read_excel(file_path, engine="openpyxl")
        except Exception as exc:
            raise RuntimeError(f"Failed to parse Excel file: {exc}") from exc

        # Capture the default sheet name
        try:
            wb = openpyxl.load_workbook(file_path, read_only=True)
            sheet_name = wb.sheetnames[0] if wb.sheetnames else None
            wb.close()
        except Exception:
            sheet_name = None

        return self._dataframe_to_tabular(df, file_type="excel", sheet_name=sheet_name)

    @staticmethod
    def _dataframe_to_tabular(
        df: "pd.DataFrame",
        file_type: str,
        sheet_name: str | None = None,
    ) -> TabularData:
        """Convert a pandas DataFrame to TabularData."""
        # Ensure headers are strings
        headers = [str(col) for col in df.columns.tolist()]

        # Convert to records, replacing NaN with None
        rows: list[dict[str, Any]] = []
        for _, row in df.iterrows():
            rows.append(
                {h: (None if pd.isna(row[h]) else row[h]) for h in headers}
            )

        return TabularData(
            headers=headers,
            rows=rows,
            row_count=len(rows),
            file_type=file_type,
            sheet_name=sheet_name,
        )
