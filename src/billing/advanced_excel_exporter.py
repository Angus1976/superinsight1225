"""
Advanced Excel Exporter for SuperInsight Platform.

Extends base Excel exporter with advanced features:
- Multi-sheet workbooks
- Charts and visualizations
- Pivot tables
- Custom templates
- Batch export
"""

from datetime import datetime, date
from typing import Dict, List, Optional, Any, Union
from pathlib import Path
from decimal import Decimal
import logging
import io

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, NamedStyle
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import BarChart, PieChart, LineChart, Reference
from openpyxl.chart.label import DataLabelList

from src.billing.excel_exporter import (
    BillingExcelExporter, ExportFormat, ExportTemplate, 
    ExportPermission, ExportStatus, ExportJob
)
from src.billing.models import BillingRecord, Bill, BillingReport

logger = logging.getLogger(__name__)


class ChartType(str):
    """Chart type enumeration."""
    BAR = "bar"
    PIE = "pie"
    LINE = "line"
    STACKED_BAR = "stacked_bar"


class TemplateManager:
    """
    Template management for Excel exports.
    
    Manages custom export templates with predefined styles and layouts.
    """
    
    def __init__(self):
        self.templates: Dict[str, Dict[str, Any]] = {}
        self._init_default_templates()
    
    def _init_default_templates(self):
        """Initialize default templates."""
        # Standard billing template
        self.templates['billing_standard'] = {
            'name': 'Standard Billing Report',
            'sheets': ['Summary', 'Details', 'User Breakdown'],
            'styles': {
                'header': {
                    'font': Font(bold=True, color="FFFFFF", size=12),
                    'fill': PatternFill(start_color="366092", end_color="366092", fill_type="solid"),
                    'alignment': Alignment(horizontal="center", vertical="center")
                },
                'data': {
                    'font': Font(size=10),
                    'alignment': Alignment(horizontal="left")
                },
                'currency': {
                    'number_format': '¥#,##0.00'
                },
                'percentage': {
                    'number_format': '0.00%'
                }
            },
            'charts': ['monthly_trend', 'user_distribution']
        }
        
        # Detailed analysis template
        self.templates['billing_detailed'] = {
            'name': 'Detailed Billing Analysis',
            'sheets': ['Executive Summary', 'Billing Details', 'User Analysis', 
                      'Project Analysis', 'Trend Analysis', 'Charts'],
            'styles': self.templates['billing_standard']['styles'],
            'charts': ['monthly_trend', 'user_distribution', 'project_breakdown', 'quality_correlation']
        }
        
        # Reward report template
        self.templates['reward_report'] = {
            'name': 'Reward Distribution Report',
            'sheets': ['Summary', 'Individual Rewards', 'Bonus Breakdown', 'Payout Schedule'],
            'styles': self.templates['billing_standard']['styles'],
            'charts': ['reward_distribution', 'bonus_types']
        }
    
    def get_template(self, template_name: str) -> Optional[Dict[str, Any]]:
        """Get a template by name."""
        return self.templates.get(template_name)
    
    def add_template(self, name: str, config: Dict[str, Any]) -> None:
        """Add a custom template."""
        self.templates[name] = config
        logger.info(f"Added template: {name}")
    
    def list_templates(self) -> List[Dict[str, str]]:
        """List available templates."""
        return [
            {'name': name, 'description': config.get('name', name)}
            for name, config in self.templates.items()
        ]


class ChartGenerator:
    """
    Chart generation for Excel exports.
    
    Creates various chart types for data visualization.
    """
    
    def __init__(self):
        self.default_colors = [
            "366092", "4F81BD", "C0504D", "9BBB59", 
            "8064A2", "4BACC6", "F79646", "2C4D75"
        ]
    
    def create_bar_chart(self, worksheet, data_range: str, 
                        categories_range: str, title: str,
                        position: str = "E2") -> BarChart:
        """Create a bar chart."""
        chart = BarChart()
        chart.type = "col"
        chart.style = 10
        chart.title = title
        
        data = Reference(worksheet, range_string=data_range)
        categories = Reference(worksheet, range_string=categories_range)
        
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)
        chart.shape = 4
        
        worksheet.add_chart(chart, position)
        return chart
    
    def create_pie_chart(self, worksheet, data_range: str,
                        categories_range: str, title: str,
                        position: str = "E2") -> PieChart:
        """Create a pie chart."""
        chart = PieChart()
        chart.title = title
        
        data = Reference(worksheet, range_string=data_range)
        categories = Reference(worksheet, range_string=categories_range)
        
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)
        
        # Add data labels
        chart.dataLabels = DataLabelList()
        chart.dataLabels.showPercent = True
        chart.dataLabels.showVal = False
        
        worksheet.add_chart(chart, position)
        return chart
    
    def create_line_chart(self, worksheet, data_range: str,
                         categories_range: str, title: str,
                         position: str = "E2") -> LineChart:
        """Create a line chart."""
        chart = LineChart()
        chart.title = title
        chart.style = 10
        
        data = Reference(worksheet, range_string=data_range)
        categories = Reference(worksheet, range_string=categories_range)
        
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)
        
        worksheet.add_chart(chart, position)
        return chart
    
    def add_trend_chart(self, worksheet, df: pd.DataFrame,
                       date_col: str, value_col: str,
                       title: str, position: str = "E2"):
        """Add a trend chart from DataFrame."""
        # Write data to worksheet
        start_row = worksheet.max_row + 2
        
        for i, (idx, row) in enumerate(df.iterrows()):
            worksheet.cell(row=start_row + i, column=1, value=row[date_col])
            worksheet.cell(row=start_row + i, column=2, value=row[value_col])
        
        end_row = start_row + len(df) - 1
        
        # Create chart
        chart = LineChart()
        chart.title = title
        
        data = Reference(worksheet, min_col=2, min_row=start_row, max_row=end_row)
        categories = Reference(worksheet, min_col=1, min_row=start_row, max_row=end_row)
        
        chart.add_data(data)
        chart.set_categories(categories)
        
        worksheet.add_chart(chart, position)


class AdvancedExcelExporter(BillingExcelExporter):
    """
    Advanced Excel exporter with enhanced features.
    
    Extends base exporter with templates, charts, and advanced formatting.
    """
    
    def __init__(self, output_dir: str = "exports"):
        super().__init__(output_dir)
        self.template_manager = TemplateManager()
        self.chart_generator = ChartGenerator()
    
    async def export_comprehensive_billing_report(
        self, 
        billing_data: Dict[str, Any],
        export_options: Dict[str, Any] = None
    ) -> Dict[str, Any]:
        """Export comprehensive billing report with all features."""
        export_options = export_options or {}
        template_name = export_options.get('template', 'billing_detailed')
        template = self.template_manager.get_template(template_name)
        
        # Create workbook
        workbook = openpyxl.Workbook()
        
        # Remove default sheet
        default_sheet = workbook.active
        workbook.remove(default_sheet)
        
        # Create sheets based on template
        if template:
            for sheet_name in template['sheets']:
                workbook.create_sheet(sheet_name)
        else:
            workbook.create_sheet('Summary')
            workbook.create_sheet('Details')
        
        # Populate sheets
        await self._create_summary_sheet(workbook, billing_data)
        await self._create_details_sheet(workbook, billing_data)
        await self._create_user_analysis_sheet(workbook, billing_data)
        
        if 'Project Analysis' in [ws.title for ws in workbook.worksheets]:
            await self._create_project_analysis_sheet(workbook, billing_data)
        
        if 'Charts' in [ws.title for ws in workbook.worksheets]:
            await self._create_charts_sheet(workbook, billing_data)
        
        # Apply formatting
        await self._apply_template_formatting(workbook, template)
        
        # Save workbook
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"billing_report_{timestamp}.xlsx"
        file_path = self.output_dir / filename
        
        workbook.save(file_path)
        
        return {
            'success': True,
            'file_path': str(file_path),
            'file_size': file_path.stat().st_size,
            'sheets_created': [ws.title for ws in workbook.worksheets],
            'export_time': datetime.now().isoformat()
        }
    
    async def _create_summary_sheet(self, workbook, billing_data: Dict[str, Any]):
        """Create summary sheet."""
        ws = workbook['Summary'] if 'Summary' in workbook.sheetnames else workbook.create_sheet('Summary')
        
        # Title
        ws['A1'] = "计费报告汇总"
        ws['A1'].font = Font(bold=True, size=16)
        ws.merge_cells('A1:D1')
        
        # Summary data
        summary = billing_data.get('summary', {})
        
        summary_data = [
            ['指标', '值'],
            ['报告期间', billing_data.get('period', 'N/A')],
            ['总标注数', summary.get('total_annotations', 0)],
            ['总工时(小时)', f"{summary.get('total_hours', 0):.2f}"],
            ['总费用', f"¥{summary.get('total_cost', 0):.2f}"],
            ['平均每小时费用', f"¥{summary.get('avg_hourly_cost', 0):.2f}"],
            ['用户数', summary.get('user_count', 0)],
            ['项目数', summary.get('project_count', 0)]
        ]
        
        for row_idx, row_data in enumerate(summary_data, start=3):
            for col_idx, value in enumerate(row_data, start=1):
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
    
    async def _create_details_sheet(self, workbook, billing_data: Dict[str, Any]):
        """Create details sheet."""
        sheet_name = 'Billing Details' if 'Billing Details' in workbook.sheetnames else 'Details'
        if sheet_name not in workbook.sheetnames:
            ws = workbook.create_sheet(sheet_name)
        else:
            ws = workbook[sheet_name]
        
        records = billing_data.get('records', [])
        
        if not records:
            ws['A1'] = "无计费记录"
            return
        
        # Headers
        headers = ['记录ID', '用户ID', '任务ID', '标注数', '工时(小时)', '费用', '日期']
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        # Data rows
        for row_idx, record in enumerate(records, start=2):
            ws.cell(row=row_idx, column=1, value=str(record.get('id', '')))
            ws.cell(row=row_idx, column=2, value=record.get('user_id', ''))
            ws.cell(row=row_idx, column=3, value=str(record.get('task_id', '')))
            ws.cell(row=row_idx, column=4, value=record.get('annotation_count', 0))
            ws.cell(row=row_idx, column=5, value=record.get('time_spent', 0) / 3600)
            ws.cell(row=row_idx, column=6, value=record.get('cost', 0))
            ws.cell(row=row_idx, column=7, value=record.get('billing_date', ''))
            
            # Format currency
            ws.cell(row=row_idx, column=6).number_format = '¥#,##0.00'
    
    async def _create_user_analysis_sheet(self, workbook, billing_data: Dict[str, Any]):
        """Create user analysis sheet."""
        sheet_name = 'User Analysis' if 'User Analysis' in workbook.sheetnames else 'User Breakdown'
        if sheet_name not in workbook.sheetnames:
            ws = workbook.create_sheet(sheet_name)
        else:
            ws = workbook[sheet_name]
        
        user_breakdown = billing_data.get('user_breakdown', {})
        
        if not user_breakdown:
            ws['A1'] = "无用户数据"
            return
        
        # Headers
        headers = ['用户ID', '标注数', '工时(小时)', '费用', '平均每小时费用', '效率评分']
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        # Data rows
        for row_idx, (user_id, data) in enumerate(user_breakdown.items(), start=2):
            hours = data.get('time_spent', 0) / 3600
            cost = data.get('cost', 0)
            
            ws.cell(row=row_idx, column=1, value=user_id)
            ws.cell(row=row_idx, column=2, value=data.get('annotations', 0))
            ws.cell(row=row_idx, column=3, value=hours)
            ws.cell(row=row_idx, column=4, value=cost)
            ws.cell(row=row_idx, column=5, value=cost / hours if hours > 0 else 0)
            ws.cell(row=row_idx, column=6, value=data.get('efficiency_score', 0))
            
            # Format
            ws.cell(row=row_idx, column=4).number_format = '¥#,##0.00'
            ws.cell(row=row_idx, column=5).number_format = '¥#,##0.00'
        
        # Add chart if enough data
        if len(user_breakdown) >= 2:
            self.chart_generator.create_bar_chart(
                ws,
                data_range=f"D2:D{len(user_breakdown)+1}",
                categories_range=f"A2:A{len(user_breakdown)+1}",
                title="用户费用分布",
                position="H2"
            )
    
    async def _create_project_analysis_sheet(self, workbook, billing_data: Dict[str, Any]):
        """Create project analysis sheet."""
        ws = workbook['Project Analysis']
        
        project_breakdown = billing_data.get('project_breakdown', {})
        
        if not project_breakdown:
            ws['A1'] = "无项目数据"
            return
        
        # Headers
        headers = ['项目ID', '项目名称', '标注数', '工时(小时)', '费用', '完成率']
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        # Data rows
        for row_idx, (project_id, data) in enumerate(project_breakdown.items(), start=2):
            ws.cell(row=row_idx, column=1, value=project_id)
            ws.cell(row=row_idx, column=2, value=data.get('name', ''))
            ws.cell(row=row_idx, column=3, value=data.get('annotations', 0))
            ws.cell(row=row_idx, column=4, value=data.get('hours', 0))
            ws.cell(row=row_idx, column=5, value=data.get('cost', 0))
            ws.cell(row=row_idx, column=6, value=data.get('completion_rate', 0))
            
            ws.cell(row=row_idx, column=5).number_format = '¥#,##0.00'
            ws.cell(row=row_idx, column=6).number_format = '0.0%'
    
    async def _create_charts_sheet(self, workbook, billing_data: Dict[str, Any]):
        """Create charts sheet."""
        ws = workbook['Charts']
        
        ws['A1'] = "数据可视化"
        ws['A1'].font = Font(bold=True, size=14)
        
        # Add placeholder for charts
        ws['A3'] = "图表将基于数据自动生成"
    
    async def _apply_template_formatting(self, workbook, template: Dict[str, Any]):
        """Apply template formatting to workbook."""
        if not template:
            return
        
        styles = template.get('styles', {})
        
        for ws in workbook.worksheets:
            # Apply header style to first row
            if 'header' in styles:
                header_style = styles['header']
                for cell in ws[1]:
                    if cell.value:
                        cell.font = header_style.get('font', Font(bold=True))
                        cell.fill = header_style.get('fill', PatternFill())
                        cell.alignment = header_style.get('alignment', Alignment())
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                ws.column_dimensions[column_letter].width = min(max_length + 2, 50)


def get_advanced_excel_exporter(output_dir: str = "exports") -> AdvancedExcelExporter:
    """Get advanced Excel exporter instance."""
    return AdvancedExcelExporter(output_dir)
