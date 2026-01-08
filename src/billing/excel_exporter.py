"""
Excel export functionality for billing system.

Provides multi-format export capabilities with customizable templates,
batch processing, and permission controls.
"""

import os
import io
from datetime import datetime, date
from typing import Dict, Any, List, Optional, Union, BinaryIO
from pathlib import Path
from enum import Enum
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import BarChart, Reference
import xlsxwriter
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.lib.units import inch
import csv
from uuid import UUID
from decimal import Decimal

from src.billing.models import BillingRecord, Bill, BillingReport
from src.billing.invoice_generator import DetailedInvoiceGenerator


class ExportFormat(str, Enum):
    """Export format enumeration."""
    EXCEL = "excel"
    PDF = "pdf"
    CSV = "csv"
    JSON = "json"


class ExportTemplate(str, Enum):
    """Export template types."""
    STANDARD = "standard"
    DETAILED = "detailed"
    SUMMARY = "summary"
    FINANCIAL = "financial"
    AUDIT = "audit"
    CUSTOM = "custom"


class ExportPermission(str, Enum):
    """Export permission levels."""
    FULL_ACCESS = "full_access"
    FINANCIAL_ONLY = "financial_only"
    SUMMARY_ONLY = "summary_only"
    NO_ACCESS = "no_access"


class ExportStatus(str, Enum):
    """Export job status."""
    PENDING = "pending"
    PROCESSING = "processing"
    COMPLETED = "completed"
    FAILED = "failed"
    CANCELLED = "cancelled"


class ExportJob:
    """Export job tracking."""
    
    def __init__(self, job_id: str, user_id: str, export_type: str, 
                 format_type: ExportFormat, template: ExportTemplate):
        self.job_id = job_id
        self.user_id = user_id
        self.export_type = export_type
        self.format_type = format_type
        self.template = template
        self.status = ExportStatus.PENDING
        self.created_at = datetime.now()
        self.started_at: Optional[datetime] = None
        self.completed_at: Optional[datetime] = None
        self.file_path: Optional[str] = None
        self.error_message: Optional[str] = None
        self.progress = 0
        
    def start(self):
        """Mark job as started."""
        self.status = ExportStatus.PROCESSING
        self.started_at = datetime.now()
        
    def complete(self, file_path: str):
        """Mark job as completed."""
        self.status = ExportStatus.COMPLETED
        self.completed_at = datetime.now()
        self.file_path = file_path
        self.progress = 100
        
    def fail(self, error_message: str):
        """Mark job as failed."""
        self.status = ExportStatus.FAILED
        self.completed_at = datetime.now()
        self.error_message = error_message
        
    def update_progress(self, progress: int):
        """Update job progress."""
        self.progress = min(100, max(0, progress))


class BillingExcelExporter:
    """
    Excel exporter for billing data with advanced formatting and templates.
    """
    
    def __init__(self, output_dir: str = "exports"):
        """Initialize the exporter."""
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(exist_ok=True)
        self.invoice_generator = DetailedInvoiceGenerator()
        self.export_jobs: Dict[str, ExportJob] = {}
        
        # Permission matrix
        self.permission_matrix = {
            ExportPermission.FULL_ACCESS: [
                "billing_records", "invoices", "reports", "audit_logs", "financial_details"
            ],
            ExportPermission.FINANCIAL_ONLY: [
                "invoices", "financial_details"
            ],
            ExportPermission.SUMMARY_ONLY: [
                "reports"
            ],
            ExportPermission.NO_ACCESS: []
        }
    
    def check_export_permission(self, user_id: str, export_type: str, 
                              permission: ExportPermission) -> bool:
        """
        Check if user has permission to export specific data type.
        
        Args:
            user_id: User identifier
            export_type: Type of data to export
            permission: User's permission level
            
        Returns:
            True if user has permission
        """
        allowed_types = self.permission_matrix.get(permission, [])
        return export_type in allowed_types
    
    def create_export_job(self, user_id: str, export_type: str, 
                         format_type: ExportFormat, template: ExportTemplate,
                         permission: ExportPermission) -> Optional[str]:
        """
        Create a new export job.
        
        Args:
            user_id: User requesting export
            export_type: Type of data to export
            format_type: Export format
            template: Export template
            permission: User's permission level
            
        Returns:
            Job ID if created successfully, None otherwise
        """
        # Check permissions
        if not self.check_export_permission(user_id, export_type, permission):
            return None
        
        # Generate job ID
        job_id = f"export_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{user_id}"
        
        # Create job
        job = ExportJob(job_id, user_id, export_type, format_type, template)
        self.export_jobs[job_id] = job
        
        return job_id
    
    def get_export_job_status(self, job_id: str) -> Optional[Dict[str, Any]]:
        """
        Get export job status.
        
        Args:
            job_id: Job identifier
            
        Returns:
            Job status information
        """
        job = self.export_jobs.get(job_id)
        if not job:
            return None
        
        return {
            "job_id": job.job_id,
            "user_id": job.user_id,
            "export_type": job.export_type,
            "format_type": job.format_type.value,
            "template": job.template.value,
            "status": job.status.value,
            "progress": job.progress,
            "created_at": job.created_at.isoformat(),
            "started_at": job.started_at.isoformat() if job.started_at else None,
            "completed_at": job.completed_at.isoformat() if job.completed_at else None,
            "file_path": job.file_path,
            "error_message": job.error_message
        }
    
    def export_billing_records_excel(self, billing_records: List[BillingRecord],
                                   template: ExportTemplate = ExportTemplate.STANDARD,
                                   filename: Optional[str] = None) -> str:
        """
        Export billing records to Excel format.
        
        Args:
            billing_records: List of billing records
            template: Export template to use
            filename: Optional custom filename
            
        Returns:
            Path to generated Excel file
        """
        if not filename:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"billing_records_{timestamp}.xlsx"
        
        file_path = self.output_dir / filename
        
        # Convert records to DataFrame
        data = []
        for record in billing_records:
            data.append({
                "记录ID": str(record.id),
                "租户ID": record.tenant_id,
                "用户ID": record.user_id,
                "任务ID": str(record.task_id) if record.task_id else "",
                "标注数量": record.annotation_count,
                "工时(秒)": record.time_spent,
                "工时(小时)": round(record.time_spent / 3600, 2),
                "费用": float(record.cost),
                "计费日期": record.billing_date.strftime("%Y-%m-%d"),
                "创建时间": record.created_at.strftime("%Y-%m-%d %H:%M:%S")
            })
        
        df = pd.DataFrame(data)
        
        # Create Excel file with formatting
        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='计费记录', index=False)
            
            # Get workbook and worksheet
            workbook = writer.book
            worksheet = writer.sheets['计费记录']
            
            # Apply formatting based on template
            if template == ExportTemplate.DETAILED:
                self._apply_detailed_formatting(workbook, worksheet, df)
            elif template == ExportTemplate.SUMMARY:
                self._apply_summary_formatting(workbook, worksheet, df)
            elif template == ExportTemplate.FINANCIAL:
                self._apply_financial_formatting(workbook, worksheet, df)
            else:  # STANDARD
                self._apply_standard_formatting(workbook, worksheet, df)
            
            # Add summary sheet if detailed template
            if template == ExportTemplate.DETAILED:
                self._add_summary_sheet(writer, billing_records)
        
        return str(file_path)
    
    def export_invoice_excel(self, invoice_data: Dict[str, Any],
                           template: ExportTemplate = ExportTemplate.STANDARD,
                           filename: Optional[str] = None) -> str:
        """
        Export invoice to Excel format.
        
        Args:
            invoice_data: Invoice data from DetailedInvoiceGenerator
            template: Export template to use
            filename: Optional custom filename
            
        Returns:
            Path to generated Excel file
        """
        if not filename:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"invoice_{invoice_data['id']}_{timestamp}.xlsx"
        
        file_path = self.output_dir / filename
        
        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            # Invoice summary sheet
            self._create_invoice_summary_sheet(writer, invoice_data)
            
            # Project breakdown sheet
            if invoice_data.get('project_breakdown'):
                self._create_project_breakdown_sheet(writer, invoice_data['project_breakdown'])
            
            # User breakdown sheet
            if invoice_data.get('user_breakdown'):
                self._create_user_breakdown_sheet(writer, invoice_data['user_breakdown'])
            
            # Financial details sheet
            if template in [ExportTemplate.DETAILED, ExportTemplate.FINANCIAL]:
                self._create_financial_details_sheet(writer, invoice_data)
        
        return str(file_path)
    
    def export_to_pdf(self, data: Dict[str, Any], 
                     template: ExportTemplate = ExportTemplate.STANDARD,
                     filename: Optional[str] = None) -> str:
        """
        Export data to PDF format.
        
        Args:
            data: Data to export
            template: Export template to use
            filename: Optional custom filename
            
        Returns:
            Path to generated PDF file
        """
        if not filename:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"export_{timestamp}.pdf"
        
        file_path = self.output_dir / filename
        
        # Create PDF document
        doc = SimpleDocTemplate(str(file_path), pagesize=A4)
        story = []
        styles = getSampleStyleSheet()
        
        # Title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            spaceAfter=30,
            alignment=1  # Center alignment
        )
        story.append(Paragraph("计费报告", title_style))
        story.append(Spacer(1, 12))
        
        # Add content based on data type
        if 'summary' in data:  # Invoice data
            self._add_invoice_pdf_content(story, data, styles)
        elif isinstance(data, list):  # Billing records
            self._add_billing_records_pdf_content(story, data, styles)
        
        # Build PDF
        doc.build(story)
        
        return str(file_path)
    
    def export_to_csv(self, data: Union[List[BillingRecord], Dict[str, Any]],
                     filename: Optional[str] = None) -> str:
        """
        Export data to CSV format.
        
        Args:
            data: Data to export
            filename: Optional custom filename
            
        Returns:
            Path to generated CSV file
        """
        if not filename:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"export_{timestamp}.csv"
        
        file_path = self.output_dir / filename
        
        with open(file_path, 'w', newline='', encoding='utf-8') as csvfile:
            if isinstance(data, list) and data and isinstance(data[0], BillingRecord):
                # Export billing records
                fieldnames = [
                    '记录ID', '租户ID', '用户ID', '任务ID', '标注数量', 
                    '工时(秒)', '费用', '计费日期', '创建时间'
                ]
                writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
                writer.writeheader()
                
                for record in data:
                    writer.writerow({
                        '记录ID': str(record.id),
                        '租户ID': record.tenant_id,
                        '用户ID': record.user_id,
                        '任务ID': str(record.task_id) if record.task_id else '',
                        '标注数量': record.annotation_count,
                        '工时(秒)': record.time_spent,
                        '费用': float(record.cost),
                        '计费日期': record.billing_date.strftime('%Y-%m-%d'),
                        '创建时间': record.created_at.strftime('%Y-%m-%d %H:%M:%S')
                    })
            
            elif isinstance(data, dict) and 'summary' in data:
                # Export invoice summary
                fieldnames = ['项目', '值']
                writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
                writer.writeheader()
                
                summary = data['summary']
                writer.writerow({'项目': '总标注数', '值': summary['total_annotations']})
                writer.writerow({'项目': '总工时(秒)', '值': summary['total_time_spent']})
                writer.writerow({'项目': '基础费用', '值': summary['base_total']})
                writer.writerow({'项目': '调整后费用', '值': summary['adjusted_subtotal']})
                writer.writerow({'项目': '最终费用', '值': summary['final_total']})
        
        return str(file_path)
    
    def batch_export(self, export_requests: List[Dict[str, Any]],
                    user_id: str, permission: ExportPermission) -> List[str]:
        """
        Process multiple export requests in batch.
        
        Args:
            export_requests: List of export request configurations
            user_id: User requesting exports
            permission: User's permission level
            
        Returns:
            List of job IDs created
        """
        job_ids = []
        
        for request in export_requests:
            job_id = self.create_export_job(
                user_id=user_id,
                export_type=request.get('export_type', 'billing_records'),
                format_type=ExportFormat(request.get('format', 'excel')),
                template=ExportTemplate(request.get('template', 'standard')),
                permission=permission
            )
            
            if job_id:
                job_ids.append(job_id)
        
        return job_ids
    
    def schedule_export(self, job_id: str, data: Any) -> bool:
        """
        Execute scheduled export job.
        
        Args:
            job_id: Job identifier
            data: Data to export
            
        Returns:
            True if export completed successfully
        """
        job = self.export_jobs.get(job_id)
        if not job:
            return False
        
        try:
            job.start()
            
            # Generate filename
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"{job.export_type}_{timestamp}"
            
            # Export based on format
            if job.format_type == ExportFormat.EXCEL:
                if job.export_type == "billing_records":
                    file_path = self.export_billing_records_excel(
                        data, job.template, f"{filename}.xlsx"
                    )
                elif job.export_type == "invoices":
                    file_path = self.export_invoice_excel(
                        data, job.template, f"{filename}.xlsx"
                    )
                else:
                    raise ValueError(f"Unsupported export type: {job.export_type}")
                    
            elif job.format_type == ExportFormat.PDF:
                file_path = self.export_to_pdf(data, job.template, f"{filename}.pdf")
                
            elif job.format_type == ExportFormat.CSV:
                file_path = self.export_to_csv(data, f"{filename}.csv")
                
            else:
                raise ValueError(f"Unsupported format: {job.format_type}")
            
            job.complete(file_path)
            return True
            
        except Exception as e:
            job.fail(str(e))
            return False
    
    def _apply_standard_formatting(self, workbook, worksheet, df):
        """Apply standard formatting to Excel worksheet."""
        # Header formatting
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        
        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    def _apply_detailed_formatting(self, workbook, worksheet, df):
        """Apply detailed formatting with borders and alternating colors."""
        self._apply_standard_formatting(workbook, worksheet, df)
        
        # Add borders
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Alternating row colors
        light_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        
        for row_num, row in enumerate(worksheet.iter_rows(min_row=2), start=2):
            for cell in row:
                cell.border = thin_border
                if row_num % 2 == 0:
                    cell.fill = light_fill
    
    def _apply_summary_formatting(self, workbook, worksheet, df):
        """Apply summary formatting with totals."""
        self._apply_standard_formatting(workbook, worksheet, df)
        
        # Add total row
        last_row = len(df) + 2
        worksheet[f'E{last_row}'] = "总计:"
        worksheet[f'F{last_row}'] = f"=SUM(F2:F{last_row-1})"  # Sum time
        worksheet[f'H{last_row}'] = f"=SUM(H2:H{last_row-1})"  # Sum cost
        
        # Format total row
        total_font = Font(bold=True)
        for col in ['E', 'F', 'H']:
            worksheet[f'{col}{last_row}'].font = total_font
    
    def _apply_financial_formatting(self, workbook, worksheet, df):
        """Apply financial formatting with currency symbols."""
        self._apply_detailed_formatting(workbook, worksheet, df)
        
        # Format currency columns
        for row in worksheet.iter_rows(min_row=2, min_col=8, max_col=8):  # Cost column
            for cell in row:
                cell.number_format = '¥#,##0.00'
    
    def _add_summary_sheet(self, writer, billing_records: List[BillingRecord]):
        """Add summary sheet to Excel workbook."""
        # Calculate summary statistics
        total_records = len(billing_records)
        total_annotations = sum(r.annotation_count for r in billing_records)
        total_time = sum(r.time_spent for r in billing_records)
        total_cost = sum(r.cost for r in billing_records)
        
        # User statistics
        user_stats = {}
        for record in billing_records:
            if record.user_id not in user_stats:
                user_stats[record.user_id] = {
                    'annotations': 0, 'time': 0, 'cost': Decimal('0')
                }
            user_stats[record.user_id]['annotations'] += record.annotation_count
            user_stats[record.user_id]['time'] += record.time_spent
            user_stats[record.user_id]['cost'] += record.cost
        
        # Create summary DataFrame
        summary_data = [
            ['总记录数', total_records],
            ['总标注数', total_annotations],
            ['总工时(小时)', round(total_time / 3600, 2)],
            ['总费用', float(total_cost)],
            ['平均每条记录费用', float(total_cost / total_records) if total_records > 0 else 0],
            ['平均每标注费用', float(total_cost / total_annotations) if total_annotations > 0 else 0]
        ]
        
        summary_df = pd.DataFrame(summary_data, columns=['指标', '值'])
        summary_df.to_excel(writer, sheet_name='汇总统计', index=False)
        
        # User breakdown
        user_data = []
        for user_id, stats in user_stats.items():
            user_data.append({
                '用户ID': user_id,
                '标注数': stats['annotations'],
                '工时(小时)': round(stats['time'] / 3600, 2),
                '费用': float(stats['cost'])
            })
        
        user_df = pd.DataFrame(user_data)
        user_df.to_excel(writer, sheet_name='用户统计', index=False)
    
    def _create_invoice_summary_sheet(self, writer, invoice_data: Dict[str, Any]):
        """Create invoice summary sheet."""
        summary = invoice_data['summary']
        
        summary_data = [
            ['发票ID', invoice_data['id']],
            ['租户ID', invoice_data['tenant_id']],
            ['计费周期', invoice_data['billing_period']],
            ['生成时间', invoice_data['generated_at']],
            ['', ''],
            ['总标注数', summary['total_annotations']],
            ['总工时(秒)', summary['total_time_spent']],
            ['基础费用', summary['base_total']],
            ['质量调整后', summary['adjusted_subtotal']],
            ['折扣后', summary['discounted_subtotal']],
            ['税费', summary['tax_amount']],
            ['最终总额', summary['final_total']]
        ]
        
        df = pd.DataFrame(summary_data, columns=['项目', '值'])
        df.to_excel(writer, sheet_name='发票汇总', index=False)
    
    def _create_project_breakdown_sheet(self, writer, project_breakdown: List[Dict[str, Any]]):
        """Create project breakdown sheet."""
        data = []
        for project in project_breakdown:
            data.append({
                '项目ID': project['project_id'],
                '项目名称': project['project_name'],
                '标注数': project['total_annotations'],
                '工时(秒)': project['total_time_spent'],
                '基础费用': project['base_cost'],
                '调整后费用': project['adjusted_cost']
            })
        
        df = pd.DataFrame(data)
        df.to_excel(writer, sheet_name='项目明细', index=False)
    
    def _create_user_breakdown_sheet(self, writer, user_breakdown: List[Dict[str, Any]]):
        """Create user breakdown sheet."""
        data = []
        for user in user_breakdown:
            data.append({
                '用户ID': user['user_id'],
                '用户名称': user['user_name'],
                '标注数': user['total_annotations'],
                '工时(秒)': user['total_time_spent'],
                '基础费用': user['base_cost'],
                '调整后费用': user['adjusted_cost']
            })
        
        df = pd.DataFrame(data)
        df.to_excel(writer, sheet_name='用户明细', index=False)
    
    def _create_financial_details_sheet(self, writer, invoice_data: Dict[str, Any]):
        """Create financial details sheet."""
        data = []
        
        # Add discount details
        if invoice_data.get('discounts'):
            for discount in invoice_data['discounts']:
                data.append({
                    '类型': '折扣',
                    '描述': discount['description'],
                    '百分比': f"{discount['percentage']}%",
                    '金额': discount['amount']
                })
        
        # Add tax details
        if invoice_data.get('tax'):
            tax = invoice_data['tax']
            data.append({
                '类型': '税费',
                '描述': tax['description'],
                '百分比': f"{tax['rate']}%",
                '金额': tax['amount']
            })
        
        if data:
            df = pd.DataFrame(data)
            df.to_excel(writer, sheet_name='财务明细', index=False)
    
    def _add_invoice_pdf_content(self, story, invoice_data: Dict[str, Any], styles):
        """Add invoice content to PDF story."""
        # Invoice header
        story.append(Paragraph(f"发票编号: {invoice_data['id']}", styles['Normal']))
        story.append(Paragraph(f"租户: {invoice_data['tenant_id']}", styles['Normal']))
        story.append(Paragraph(f"计费周期: {invoice_data['billing_period']}", styles['Normal']))
        story.append(Spacer(1, 12))
        
        # Summary table
        summary = invoice_data['summary']
        summary_data = [
            ['项目', '值'],
            ['总标注数', str(summary['total_annotations'])],
            ['总工时(秒)', str(summary['total_time_spent'])],
            ['基础费用', f"¥{summary['base_total']:.2f}"],
            ['最终总额', f"¥{summary['final_total']:.2f}"]
        ]
        
        summary_table = Table(summary_data)
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 14),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        story.append(summary_table)
    
    def _add_billing_records_pdf_content(self, story, billing_records: List[BillingRecord], styles):
        """Add billing records content to PDF story."""
        # Records table
        data = [['用户ID', '标注数', '工时(小时)', '费用']]
        
        for record in billing_records[:20]:  # Limit to first 20 records for PDF
            data.append([
                record.user_id,
                str(record.annotation_count),
                f"{record.time_spent / 3600:.2f}",
                f"¥{float(record.cost):.2f}"
            ])
        
        table = Table(data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        story.append(table)
        
        if len(billing_records) > 20:
            story.append(Spacer(1, 12))
            story.append(Paragraph(f"注: 仅显示前20条记录，总共{len(billing_records)}条记录", styles['Normal']))


class ExportScheduler:
    """
    Export scheduler for handling timed and batch exports.
    """
    
    def __init__(self, exporter: BillingExcelExporter):
        """Initialize the scheduler."""
        self.exporter = exporter
        self.scheduled_jobs: Dict[str, Dict[str, Any]] = {}
    
    def schedule_daily_export(self, tenant_id: str, export_config: Dict[str, Any]) -> str:
        """
        Schedule daily export for a tenant.
        
        Args:
            tenant_id: Tenant identifier
            export_config: Export configuration
            
        Returns:
            Schedule ID
        """
        schedule_id = f"daily_{tenant_id}_{datetime.now().strftime('%Y%m%d')}"
        
        self.scheduled_jobs[schedule_id] = {
            "tenant_id": tenant_id,
            "schedule_type": "daily",
            "config": export_config,
            "created_at": datetime.now(),
            "next_run": datetime.now().replace(hour=0, minute=0, second=0) + pd.Timedelta(days=1)
        }
        
        return schedule_id
    
    def schedule_weekly_export(self, tenant_id: str, export_config: Dict[str, Any]) -> str:
        """
        Schedule weekly export for a tenant.
        
        Args:
            tenant_id: Tenant identifier
            export_config: Export configuration
            
        Returns:
            Schedule ID
        """
        schedule_id = f"weekly_{tenant_id}_{datetime.now().strftime('%Y%m%d')}"
        
        self.scheduled_jobs[schedule_id] = {
            "tenant_id": tenant_id,
            "schedule_type": "weekly",
            "config": export_config,
            "created_at": datetime.now(),
            "next_run": datetime.now() + pd.Timedelta(weeks=1)
        }
        
        return schedule_id
    
    def process_scheduled_jobs(self) -> List[str]:
        """
        Process all scheduled jobs that are due.
        
        Returns:
            List of processed job IDs
        """
        processed_jobs = []
        current_time = datetime.now()
        
        for schedule_id, job_info in self.scheduled_jobs.items():
            if current_time >= job_info["next_run"]:
                # Process the job
                # This would typically fetch data and create export job
                processed_jobs.append(schedule_id)
                
                # Update next run time
                if job_info["schedule_type"] == "daily":
                    job_info["next_run"] += pd.Timedelta(days=1)
                elif job_info["schedule_type"] == "weekly":
                    job_info["next_run"] += pd.Timedelta(weeks=1)
        
        return processed_jobs